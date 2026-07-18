from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path
from typing import Any


CONFIRMED_SHEET_NAME = "已確認可用"
IGNORED_SHEETS = {"待人工確認", "不可使用", "重複案例", "欄位完整率", "公司部門職務統計", "問題類型統計"}

REFERENCE_SECTION_FIELDS = [
    "section_1_1_work_pattern",
    "section_1_2_work_characteristics",
    "section_2_1_data_basis",
    "section_2_2_risk_analysis",
    "section_3_1_management_advice",
    "section_3_2_environment_advice",
    "section_3_3_health_guidance",
    "section_4_1_guidance_result",
    "section_4_2_work_suitability",
    "section_4_3_follow_up",
]

SEARCH_FIELDS = [
    "company",
    "industry",
    "department",
    "job_title",
    "work_shift",
    "main_work_tasks",
    "special_health_risks",
    "other_health_risks",
    "supplementary_notes",
    "other_special_hazards",
    "noise_management",
    "dust_management",
    "ionizing_radiation_management",
]

DOMAIN_KEYWORDS = [
    "久站",
    "久坐",
    "電腦",
    "文書",
    "搬運",
    "重複",
    "手部",
    "人因",
    "肌肉骨骼",
    "肩頸",
    "下背",
    "腰",
    "輪班",
    "夜間",
    "長工時",
    "疲勞",
    "高溫",
    "高氣溫",
    "戶外",
    "熱",
    "噪音",
    "聽力",
    "粉塵",
    "呼吸",
    "游離輻射",
    "輻射",
    "化學",
    "清潔劑",
    "溶劑",
    "滑倒",
    "跌倒",
    "交通",
    "駕駛",
    "重機具",
    "監督",
    "管理",
    "工程",
    "工地",
    "現場",
    "溝通",
    "協調",
    "壓力",
    "高血壓",
    "血糖",
    "血脂",
    "慢性病",
    "中高齡",
    "孕產婦",
    "貧血",
    "肝功能",
    "心理",
]

LEVEL_ALIASES = {
    "一級": "第一級",
    "二級": "第二級",
    "三級": "第三級",
    "四級": "第四級",
}


def normalize_text(value: Any) -> str:
    if isinstance(value, list):
        value = "、".join(str(item) for item in value if str(item).strip())
    text = str(value or "").strip()
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    for old, new in LEVEL_ALIASES.items():
        text = text.replace(old, new)
    return text


def split_items(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [item.strip() for item in re.split(r"[、,，;；/／\n]+", text) if item.strip()]


def build_query_text(form_data: dict) -> str:
    special_hazards = []
    if normalize_text(form_data.get("noise_health_level")):
        special_hazards.append("噪音")
    if normalize_text(form_data.get("dust_health_level")):
        special_hazards.append("粉塵")
    if normalize_text(form_data.get("radiation_health_level")):
        special_hazards.append("游離輻射")

    parts = [
        form_data.get("company", ""),
        form_data.get("industry", ""),
        form_data.get("department", ""),
        form_data.get("job_title", ""),
        form_data.get("shift", ""),
        form_data.get("work_content", ""),
        form_data.get("health_risks", ""),
        form_data.get("custom_health_risks", ""),
        form_data.get("notes", ""),
        "、".join(special_hazards),
    ]
    return normalize_text(" ".join(normalize_text(part) for part in parts if normalize_text(part)))


def tokenize(text: str) -> set[str]:
    text = normalize_text(text).lower()
    tokens = set()
    for keyword in DOMAIN_KEYWORDS:
        if keyword.lower() in text:
            tokens.add(keyword)
    tokens.update(re.findall(r"[a-z0-9]{2,}", text))
    cjk = re.sub(r"[^\u4e00-\u9fff]", "", text)
    tokens.update(cjk[i : i + 2] for i in range(max(0, len(cjk) - 1)))
    return {token for token in tokens if token.strip()}


def token_similarity(left: str, right: str) -> float:
    left_tokens = tokenize(left)
    right_tokens = tokenize(right)
    if not left_tokens or not right_tokens:
        return 0.0
    return len(left_tokens & right_tokens) / len(left_tokens | right_tokens)


def field_similarity(form_value: Any, case_value: Any) -> float:
    left = normalize_text(form_value)
    right = normalize_text(case_value)
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    if left in right or right in left:
        return 0.78
    return token_similarity(left, right)


def normalize_case(row: dict[str, Any]) -> dict[str, str]:
    case = {}
    allowed_fields = [
        "案例ID",
        "日期",
        "來源檔案",
        "company",
        "industry",
        "department",
        "job_title",
        "work_shift",
        "health_management_year",
        "health_management_level",
        "special_health_risks",
        "other_health_risks",
        "work_ability_tag",
        "main_work_tasks",
        "supplementary_notes",
        "special_hazard_year",
        "noise_management",
        "dust_management",
        "ionizing_radiation_management",
        "other_special_hazards",
        "assessment_result",
        "work_suitability_category",
        "work_suitability_detail",
        "custom_text",
        *REFERENCE_SECTION_FIELDS,
    ]
    for field in allowed_fields:
        case[field] = normalize_text(row.get(field))
    return case


@lru_cache(maxsize=2)
def load_appendix8_cases(workbook_path: str) -> tuple[dict[str, str], ...]:
    path = Path(workbook_path)
    if not path.exists():
        return tuple()

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if CONFIRMED_SHEET_NAME not in workbook.sheetnames:
        return tuple()

    worksheet = workbook[CONFIRMED_SHEET_NAME]
    headers = [normalize_text(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    cases = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
        case = normalize_case(row_data)
        if any(case.get(field) for field in REFERENCE_SECTION_FIELDS):
            cases.append(case)
    return tuple(cases)


def case_search_text(case: dict[str, str]) -> str:
    return " ".join(case.get(field, "") for field in SEARCH_FIELDS)


def special_hazard_text(form_data: dict) -> str:
    hazards = []
    if normalize_text(form_data.get("noise_health_level")):
        hazards.append("噪音")
    if normalize_text(form_data.get("dust_health_level")):
        hazards.append("粉塵")
    if normalize_text(form_data.get("radiation_health_level")):
        hazards.append("游離輻射")
    return "、".join(hazards)


def score_case(form_data: dict, case: dict[str, str]) -> float:
    weighted_scores = [
        (field_similarity(form_data.get("company"), case.get("company")), 2.0),
        (field_similarity(form_data.get("industry"), case.get("industry")), 2.0),
        (field_similarity(form_data.get("department"), case.get("department")), 3.0),
        (field_similarity(form_data.get("job_title"), case.get("job_title")), 4.0),
        (field_similarity(form_data.get("shift"), case.get("work_shift")), 2.0),
        (field_similarity(form_data.get("work_content"), case.get("main_work_tasks")), 6.0),
        (field_similarity(form_data.get("health_risks"), case.get("special_health_risks")), 4.0),
        (field_similarity(form_data.get("custom_health_risks"), case.get("other_health_risks")), 2.0),
        (field_similarity(special_hazard_text(form_data), " ".join([
            case.get("noise_management", ""),
            case.get("dust_management", ""),
            case.get("ionizing_radiation_management", ""),
            case.get("other_special_hazards", ""),
        ])), 4.0),
        (token_similarity(build_query_text(form_data), case_search_text(case)), 5.0),
    ]
    score = sum(value * weight for value, weight in weighted_scores) / sum(weight for _, weight in weighted_scores)

    company = normalize_text(form_data.get("company"))
    if company and case.get("company") and company != case.get("company"):
        score *= 0.82
    job_title = normalize_text(form_data.get("job_title"))
    if job_title and case.get("job_title") and field_similarity(job_title, case.get("job_title")) < 0.12:
        score *= 0.9
    return round(score, 4)


def find_similar_appendix8_cases(
    form_data: dict,
    workbook_path: str | Path,
    *,
    limit: int = 3,
    min_score: float = 0.22,
) -> list[dict[str, Any]]:
    cases = load_appendix8_cases(str(workbook_path))
    scored = []
    for case in cases:
        score = score_case(form_data, case)
        if score >= min_score:
            scored.append({"score": score, "case": case})
    scored.sort(key=lambda item: item["score"], reverse=True)
    return scored[:limit]


def text_has_any(text: str, keywords: list[str]) -> bool:
    clean = normalize_text(text)
    return any(keyword in clean for keyword in keywords)


def current_context_text(form_data: dict, cases: list[dict[str, Any]]) -> str:
    # Only the current form can activate a risk theme. Similar cases are used to
    # select writing direction, but their unrelated hazards must not leak into
    # this assessment.
    return build_query_text(form_data)


def build_reference_guidance(form_data: dict, similar_cases: list[dict[str, Any]]) -> dict[str, list[str]]:
    if not similar_cases:
        return {"risk": [], "management": [], "environment": [], "education": [], "follow_up": []}

    context = current_context_text(form_data, similar_cases)
    guidance = {"risk": [], "management": [], "environment": [], "education": [], "follow_up": []}

    if text_has_any(context, ["工地", "營造", "現場巡檢", "戶外巡檢", "施工", "工安巡查"]):
        guidance["risk"].append("本次作業涉及現場監督或工地管理，需留意戶外移動、工程進度壓力、臨時狀況處理及現場環境變化對疲勞與判斷力之影響。")
        guidance["management"].append("工作負荷與現場管理：建議依工程進度、巡視頻率及突發狀況處理需求，檢視工作分派、支援機制與休息安排。")
        guidance["follow_up"].append("追蹤現場監督、巡檢頻率、工時安排及突發狀況處理對疲勞與作業安全之影響。")

    if text_has_any(context, ["久坐", "電腦", "文書", "辦公室", "繪圖"]):
        guidance["risk"].append("久坐與電腦或文書作業可能增加肩頸、腰背、腕部及視覺疲勞等人因性負荷。")
        guidance["environment"].append("建議檢視座椅支撐、螢幕高度、鍵盤滑鼠距離、桌面配置及照明條件，以降低長時間電腦作業負荷。")
        guidance["education"].append("指導個案於電腦或文書作業期間定時變換姿勢，執行肩頸、腰背、手腕與視覺休息。")

    if text_has_any(context, ["久站", "站立"]):
        guidance["risk"].append("長時間站立可能增加下肢痠痛、足部不適、膝踝負荷及靜脈循環負擔。")
        guidance["environment"].append("建議檢視站立工作區地墊、工作高度、可短暫坐下休息位置及常用物品擺放。")
        guidance["education"].append("指導久站作業中進行下肢伸展、重心轉換及足部不適警訊辨識。")

    if text_has_any(context, ["搬運", "抬舉", "推拉", "重量"]):
        guidance["risk"].append("搬運、抬舉或推拉作業可能增加下背、肩頸、膝部及上肢肌肉骨骼傷害風險。")
        guidance["management"].append("搬運負荷管理：建議檢視單次重量、搬運頻率、動線距離及是否需使用推車、升降台或雙人作業。")
        guidance["education"].append("指導搬運前評估重量與路線，避免彎腰扭轉施力，必要時使用輔具或協同搬運。")

    if text_has_any(context, ["高溫", "高氣溫", "戶外", "熱", "溫差"]):
        guidance["risk"].append("高氣溫、戶外或溫差暴露可能增加熱不適、疲勞、注意力下降及心血管負荷。")
        guidance["management"].append("熱危害與負荷管理：建議依熱暴露程度、戶外停留時間及作業強度，調整巡視頻率、輪替制度與休息安排。")
        guidance["environment"].append("建議確認遮蔭休息區、通風降溫設備、補水點及熱危害警示或通報流程。")
        guidance["education"].append("指導補水、休息與熱危害警訊辨識，若出現頭暈、胸悶、極度疲倦或步態不穩應即時反映並休息。")

    if text_has_any(context, ["噪音", "聽力", "耳塞", "耳罩"]):
        guidance["risk"].append("噪音暴露可能增加聽力損失、耳鳴、溝通困難及警示聲辨識下降之風險。")
        guidance["management"].append("噪音作業管理：建議檢視噪音暴露時間、作業輪替、聽力防護具配戴及保存管理。")
        guidance["environment"].append("建議確認噪音源隔離、設備維護、警示標示及耳塞或耳罩配置是否符合現場需求。")
        guidance["education"].append("提醒正確配戴聽力防護具，留意耳鳴、聽力下降或溝通困難等警訊。")

    if text_has_any(context, ["粉塵", "呼吸", "水泥", "矽塵"]):
        guidance["risk"].append("粉塵暴露可能增加眼鼻喉刺激、咳嗽、胸悶及慢性呼吸道負荷。")
        guidance["management"].append("粉塵作業管理：建議檢視粉塵暴露時間、作業方式、局部排氣及呼吸防護具使用管理。")
        guidance["environment"].append("建議確認濕式作業、局部排氣、清掃方式、粉塵逸散控制及呼吸防護具配置。")
        guidance["education"].append("提醒正確配戴合適呼吸防護具，避免乾掃造成二次揚塵，並留意咳嗽或呼吸道刺激症狀。")

    has_management_role = text_has_any(context, ["管理", "管理職", "主管", "人員管理", "溝通協調", "行政管理", "一般管理"])
    if has_management_role:
        guidance["risk"].append("管理職需承擔決策、人員管理及溝通協調責任，可能因工作壓力、長工時、休息不足及久坐辦公，增加疲勞、心理負荷、肌肉骨骼不適及腦心血管疾病風險。")
        guidance["management"].append("工作負荷與壓力管理：建議合理分配工作與責任，建立明確溝通及支援機制，並安排適當休息。")
        guidance["education"].append("提供壓力調適、工作節奏安排、久坐辦公伸展及異常疲勞警訊辨識之指導。")
    elif text_has_any(context, ["溝通", "協調", "壓力", "心理", "跨部門"]):
        guidance["risk"].append("工作需頻繁溝通、協調或處理突發事項時，可能增加心理壓力、疲勞累積及注意力下降風險。")
        guidance["management"].append("工作負荷與壓力管理：建議合理分配工作與責任，建立明確溝通及支援機制，並安排適當休息。")
        guidance["education"].append("提供壓力調適、工作節奏安排及異常疲勞警訊辨識之指導。")

    if text_has_any(context, ["中高齡", "慢性病", "高血壓", "血糖", "血脂", "心血管"]):
        guidance["risk"].append("若合併中高齡、慢性病或心血管相關風險，需特別留意作業負荷、疲勞、熱暴露及長工時對健康耐受度之影響。")
        guidance["follow_up"].append("後續追蹤以作業負荷、休息安排、環境暴露及工作適性變化為主，必要時再評估工作調整需求。")

    return {key: dedupe_texts(value) for key, value in guidance.items()}


def dedupe_texts(items: list[str]) -> list[str]:
    seen = set()
    result = []
    for item in items:
        clean = normalize_text(item).rstrip("。；;") + "。"
        if clean and clean not in seen:
            seen.add(clean)
            result.append(clean)
    return result


def format_case_match_summary(similar_cases: list[dict[str, Any]]) -> str:
    if not similar_cases:
        return "未找到高度相似案例，已使用通用風險規則產生。"
    lines = []
    for index, item in enumerate(similar_cases, start=1):
        case = item["case"]
        labels = [
            case.get("company"),
            case.get("industry"),
            case.get("department"),
            case.get("job_title"),
            case.get("main_work_tasks"),
        ]
        label = "／".join(part for part in labels if part) or case.get("案例ID", "未命名案例")
        lines.append(f"{index}. 相似度 {item['score']:.2f}：{label}")
    return "\n".join(lines)
